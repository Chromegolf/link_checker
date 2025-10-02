##TODO
##сгенерировать отчет из какой-нибудь папки
##переименовать отчет - для удобства работы. Можно переименовать в их код из ReportService
import random
import openpyxl
import webbrowser

wb = openpyxl.load_workbook('issue.xlsx')
ws = wb.active
max_col = ws.max_column
max_row = ws.max_row


def get_issue_screen_link(ws, max_row):
    screen_links = [ws[f'X{row}'].value for row in range(5, max_row)]

    for link in random.sample(screen_links, min(3, len(screen_links))):
        if link:  # проверяем, что ячейка не пустая
            webbrowser.open(link)


def get_issue_screen_check_link(ws, max_row):
    screen_links = [ws[f'Y{row}'].value for row in range(5, max_row)]

    for link in random.sample(screen_links, len(screen_links)):
        if link:  # проверяем, что ячейка не пустая
            webbrowser.open(link)


if __name__ == '__main__':
    get_issue_screen_link(ws, max_row)
    get_issue_screen_check_link(ws, max_row)
